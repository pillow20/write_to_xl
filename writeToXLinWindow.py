import tkinter as tk
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

class DataEntryApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Ввод данных')
        self.root.geometry('500x500')
        self.entries = []
        self.create_widgets()
        self.filename = 'data.xlsx'
        # Загружаем существующий файл или создаем новый
        if os.path.exists(self.filename):
            self.workbook = load_workbook(self.filename)
        else:
            self.workbook = Workbook()

    def create_widgets(self):
        for i in range(1, 11):
            label = tk.Label(self.root, text=f'Введите текст №{i}:')
            label.pack()
            entry = tk.Entry(self.root)
            entry.pack()
            self.entries.append(entry)
        save_button = tk.Button(self.root, text='Сохранить', command=self.save_data)
        save_button.pack()

    def save_data(self):
        # Проверяем, есть ли данные для сохранения
        if any(entry.get() for entry in self.entries):
            # Создаем новый лист с уникальным именем, основанным на текущем времени
            current_time = datetime.now().strftime('%Y-%m-%d %H%M%S ')
            self.sheet = self.workbook.create_sheet(title=current_time)
            row = 1  # Начинаем запись данных с первой строки
            col = 'H'
            for entry in self.entries:
                self.sheet[f'{col}{row}'] = entry.get()
                row += 1  # Переходим на следующую строку
                entry.delete(0, tk.END)
            self.workbook.save(self.filename)
        else:
            print("Нет данных для сохранения.")

    def on_closing(self):
        # Сохраняем данные при закрытии окна, если есть что сохранять
        self.save_data()
        self.root.destroy()

# Создаем главное окно
root = tk.Tk()
app = DataEntryApp(root)
root.protocol("WM_DELETE_WINDOW", app.on_closing)  # Обработчик закрытия окна
root.mainloop()
